import glob, os, sys
import numpy as np
import codecs
from copy import deepcopy
from openpyxl import Workbook
import collections
import re
import yaml
# import colored
from termcolor import colored, cprint
import argparse
import math
#
from colorclass import Color, Windows
from terminaltables import SingleTable, DoubleTable


def decode_task(code):
    decode = "MT+"
    code_dict = {
        '0': 'amr',
        '1': 'tree',
        '2': 'ner',
        '3': 'pos',
    }
    code_list = ["AMR", "TREE", "NER", "POS"]
    for i in range(len(code)):
        if code[i] == "0":
            pass
        elif code[i] == "1":
            decode = decode + code_dict[str(i)] + "+"
        else:
            decode = decode + code_list[i] + str(code[i]) + "+"
    decode = decode.strip("+")
    return decode


def extract_info(info_dict, info_file_path, train_file_path):
    # find requested info from info file
    for line in open(info_file_path, "r").readlines():
        item_name = line.split(":")[0].strip()
        item_value = line.split(":")[1].strip().strip("\'")
        if item_name in info_dict.keys():
            info_dict[item_name] = item_value
    # rewrite some param values with the values in the training log
    with open(train_file_path, "r") as log_file:
        info_line = ""
        for i in range(20):
            line = log_file.readline()
            if "Git Commit" in line:
                info_dict["Git_Commit"] = line.split()[5]
            if "Namespace" in line:
                info_line = line
                break
        actual_line = info_line[info_line.find("(") + 1:-2]
        splitted_info = actual_line.split(", ")
        for item in splitted_info:
            key = item.split("=")[0].strip()
            value = item.split("=")[1].strip().strip("\'")
            if key in info_dict.keys():
                info_dict[key] = value
    return info_dict


def text_dict_to_string(dict):
    message = ""
    for key, value in dict.items():
        if key.startswith("SEP"):
            message += "\n\t"
        else:
            new_line = "%s: %s" % (key, value)
            if key in colored_dict.keys():
                color = colored_dict[key]
                if color == "hidden":
                    continue
                new_line = colored(new_line, color)
            message += new_line + ", "
    return message


def text_stats_to_string(min_pplx, best_epoch, last_epoch, last_step, job_info, test_bleu_list, test_pplx_list,
                         dev_bleu_list, dev_pplx_list):
    output = "\nBest dev PPLX during training: " + colored(str(min_pplx), "magenta") + ", Epoch:" + str(
        best_epoch) + ", " + str(
        last_epoch) + ", Step: " + last_step + job_info + "\n"

    output += "Test BLEU/PPLX of saved models: " + "\n"
    output += colored("%s,%s" % (test_bleu_list, test_pplx_list), "magenta") + "\n"
    output += "Dev BLEU/PPLX of saved models: " + "\n"
    output += colored("%s,%s" % (dev_bleu_list, dev_pplx_list), "magenta") + "\n"
    output += "Dev PPLX during the training: " + "\n"
    output += str(pplx_hist) + "\n"
    return output


def _list_to_string(list, element_per_line=7):
    output = ""
    for idx, val in enumerate(list):
        output += str(val)
        if (idx + 1) % element_per_line == 0:
            output += "\n"
        else:
            output += ", "
    return output.strip(", ").strip()


def table_stats_to_string(min_pplx, best_epoch, last_epoch, last_step, job_info, test_bleu_list, test_pplx_list,
                          dev_bleu_list, dev_pplx_list):
    table_data = []
    current_row = ["Job info:", str(job_info).strip()]
    table_data.append(current_row)
    current_row = ["Step (epoch):", str(last_step) + " (%s/%s)" % (str(best_epoch), str(last_epoch))]
    table_data.append(current_row)
    current_row = ["Best dev PPLX (train)", colored(str(min_pplx), "magenta")]
    table_data.append(current_row)

    current_row = ["Test BLEU of saved models: ",
                   colored("%s" % (test_bleu_list), "magenta", attrs=['bold'])]
    table_data.append(current_row)
    current_row = ["Test PPLX of saved models: ",
                   colored("%s" % (test_pplx_list), "magenta", attrs=['bold'])]
    table_data.append(current_row)
    current_row = ["Dev BLEU of saved models: ",
                   colored("%s" % (dev_bleu_list), "magenta", attrs=['bold'])]
    table_data.append(current_row)
    current_row = ["Dev PPLX during the training: ", _list_to_string(pplx_hist)]
    table_data.append(current_row)
    table_data = _reshape_1d_table(table_data, 1)
    table_instance = SingleTable(table_data, "Stats")
    table_instance.inner_heading_row_border = False
    table_instance.inner_row_border = True
    # print(table_instance.ok)
    return colored(table_instance.table, attrs=['bold'])


def _reshape_1d_table(table, cols):
    new_table = []
    total_items = len(table)
    items_per_col = math.ceil(float(total_items) / cols)
    current_col = ["", ""]
    item_counter = 0
    for key, value in table:
        current_col[0] += "\n%s" % key
        current_col[1] += "\n%s" % value
        item_counter += 1
        if item_counter == items_per_col:
            new_table.append(current_col[0].strip())
            new_table.append(current_col[1].strip())
            current_col = ["", ""]
            item_counter = 0
    if current_col[0] != "":
        new_table.append(current_col[0].strip())
        new_table.append(current_col[1].strip())
    return [new_table]


def table_dict_to_string(dict):
    message = ""

    table_data = []
    current_row = []
    for key, value in dict.items():
        if key.startswith("SEP"):
            message += "\n\t"
            current_row.append("**** %s *****"%key.split("-")[1])
            current_row.append("**********")
            table_data.append(current_row)
            current_row = []
        elif key is "Git_Commit":
            continue
        else:
            filtered_value = (value[:15] + '..') if len(value) > 15 else value
            new_line = "%s: %s" % (key, filtered_value)
            colored_key = key
            colored_value = filtered_value if filtered_value.strip() != "" else "-"
            if key in colored_dict.keys():
                color = colored_dict[key]
                if color == "hidden":
                    continue
                new_line = colored(new_line, color)
                colored_key = colored(key, color)
                colored_value = colored(filtered_value, color)

            current_row.append(colored_key)
            current_row.append(colored_value)
            message += new_line + ", "
            table_data.append(current_row)
            current_row = []

    for i in reversed(range(args.cols)):
        shaped_table_data = _reshape_1d_table(table_data, i)
        table_instance = SingleTable(shaped_table_data, "Configurations")
        table_instance.inner_heading_row_border = False
        table_instance.inner_row_border = True
        if table_instance.ok:
            break

    output = "Git commit: %s\n%s" % (dict["Git_Commit"], table_instance.table)

    return output


def find_occurrences(s, ch):
    return [i for i, letter in enumerate(s) if letter == ch]


def find_job_index(jobs, job_name):
    counter = 0
    for job in jobs:
        for info in job:
            if info == job_name:
                return counter
        counter += 1
    return -1


# Return training statistics
def train_stats(train_file_path):
    min_pplx = 100000.0
    pplx_hist = []
    best_epoch = ""
    last_epoch = 0
    last_step = ""
    with open(train_file_path) as file:
        for line in file:
            if "(Task 0) Validation perplexity" in line:
                last_epoch += 1
                splitted_line = line.split()
                val_pplx = float(splitted_line[-1])
                pplx_hist.append(val_pplx)
                if min_pplx > val_pplx:
                    min_pplx = val_pplx
                    best_epoch = last_epoch
            if "Step" in line:
                last_step = line.split()[6].strip(";")
    return min_pplx, best_epoch, last_epoch, pplx_hist, last_step


# Multiply time by ratio (required for calculating the remaining time)
def multiply_time_str(time_str, ratio):
    time_seconds = 0
    splitted_time = re.split(":|-", time_str)
    multipliers = [3600 * 24, 3600, 60, 1]
    splitted_time.reverse()
    multipliers.reverse()
    for index in range(len(splitted_time)):
        time_seconds += int(splitted_time[index]) * multipliers[index]
    new_time_seconds = int(time_seconds * ratio)
    days = new_time_seconds // (3600 * 24)
    hours = new_time_seconds // 3600 - (days * 24)
    minutes = (new_time_seconds // 60) - (days * 60 * 24) - (hours * 60)
    seconds = new_time_seconds - (days * 24 * 3600) - (hours * 3600) - (minutes * 60)

    new_date = "%02i-%02i:%02i:%02i" % (days, hours, minutes, seconds)
    return new_date


# Extract BLEU scores of saved models from the Checkpoints folder
def bleu_stats(chkpt_folder, ext="tbleu"):
    bleu_hist = []
    try:
        for dir in os.listdir(chkpt_folder):
            if dir.endswith(ext):
                bleu_file = os.path.abspath(chkpt_folder + dir)
                with open(bleu_file) as st:
                    stat_line = st.readline()
                    bleu = stat_line[6:stat_line.find(",")].strip()
                    # print(bleu)
                    bleu_hist.append(float(bleu))
    except:
        bleu_hist = []
    return bleu_hist


# Extract PPLX of saved models from the Checkpoints folder
def pplx_stats(chkpt_folder, ext=".stats"):
    pplx_hist = []
    try:
        for dir in os.listdir(chkpt_folder):
            if dir.endswith(ext):
                bleu_file = os.path.abspath(chkpt_folder + dir)
                for stat_line in open(bleu_file):
                    if stat_line.startswith("GOLD"):
                        pplx = stat_line.split()[6]
                        pplx_hist.append(float(pplx))
    except:
        pplx_hist = []
    return pplx_hist


def verify_filters(filters_dict, filled_info_dict):
    is_passed = True
    for key, value in filters_dict.items():
        if filled_info_dict[key] != value:
            is_passed = False
            break
    return is_passed


if __name__ == "__main__":

    parser = argparse.ArgumentParser(description='Experiments statistics.')
    parser.add_argument('--config', '-config', action='store', dest='config', help='Path to config file.',
                        default="configs/exp_stats.yml")
    parser.add_argument('--filters', '-filters', action='store', dest='filters', help='filters separate by comma',
                        default="")
    parser.add_argument('--folders', '-folders', action='store', dest='folders', help='folders separate by comma',
                        default="models")
    parser.add_argument('--info', '-info', action='store', dest='info',
                        help='info which should be extracted from experiments',
                        default="")
    parser.add_argument('--cols', '-cols', action='store', dest='cols',
                        help='Maximum number of table columns',
                        type=int, default=7)
    parser.add_argument('--text_view', '-text_view', action="store_true",
                        help="Switch for table or text view", default=False)
    parser.add_argument('--no_status', '-no_status', action="store_true",
                        help="No command line status", default=False)

    args = parser.parse_args()

    # Read config file
    with open(args.config, 'r') as stream:
        try:
            configs = yaml.load(stream)
        except yaml.YAMLError as exc:
            raise AssertionError("Cannot read config file. It should be in YAML format.")

    # Take union of options of config file and command-line arguments
    for option in ["info", 'filters']:
        if configs[option] is None:
            configs[option] = []
        if vars(args)[option] != "":
            for item in vars(args)[option].split(','):
                configs[option].append(item)

    # Create dictionary for requested info
    info_dict = collections.OrderedDict()
    colored_dict = collections.OrderedDict()
    for item in configs["info"]:
        if isinstance(item, dict):
            key = list(item.keys())[0]
            color = item[key]
            info_dict[key] = "Not found!"
            colored_dict[key] = color
        else:
            info_dict[item] = "Not found!"

    # Create dictionary for  filters
    filters_dict = collections.OrderedDict()
    for item in configs["filters"]:
        item_key = item.split("=")[0]
        item_value = item.split("=")[1]
        filters_dict[item_key] = item_value

    if 'epochs' not in info_dict:
        info_dict['epochs'] = "Not found!"

    # extract the name of jobs which are currently running (only for SLURM job managers)
    command = "squeue > .running_jobs"
    os.system(command)
    jobs = []
    first = True
    with open(".running_jobs") as file:
        for line in file:
            if first:
                first = False
            else:
                splitted_line = line.strip().split()
                jobs.append([splitted_line[0], splitted_line[1], splitted_line[2], splitted_line[5]])

    # Create an excel sheet
    wb = Workbook()
    ws = wb.active
    line = 1

    dict_to_string = text_dict_to_string if args.text_view else table_dict_to_string
    stats_to_string = text_stats_to_string if args.text_view else table_stats_to_string


    def _maybe_print(input):
        if not args.no_status:
            print(input )


    folder_list = args.folders.split(",")
    if folder_list[0] == "*":
        folder_list = [path for path in os.listdir(".") if os.path.isdir(path)]
    for model_folder in folder_list:
        print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
        print("folder: " + model_folder)
        print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
        for maybe_dir in sorted(os.listdir(model_folder)):
            if os.path.isdir(os.path.join(model_folder, maybe_dir)):
                dir = maybe_dir
                full_folder = model_folder + "/" + dir + "/"
                splitted_dir = dir.split("_")
                train_log_file = full_folder + "train_log"
                info_log_file = full_folder + "info"
                # Check for the train log
                if os.path.isfile(train_log_file):
                    try:
                        filled_info_dict = extract_info(deepcopy(info_dict), info_log_file, train_log_file)
                        min_pplx, best_epoch, last_epoch, pplx_hist, last_step = train_stats(train_log_file)

                        filled_info_dict["is_trained"] = "True" if str(last_epoch) == filled_info_dict[
                            "epochs"] else "False"
                        filled_info_dict["is_training"] = "False"

                        job_index = find_job_index(jobs, "m" + splitted_dir[-1])
                        job_info = "-"
                        # Check whether it is a running job
                        if job_index != -1:
                            filled_info_dict["is_training"] = "True"
                            job_info = " (" + jobs[job_index][1] + ", " + jobs[job_index][0] + ", " + jobs[job_index][
                                3] + ")"
                            # Calculate remaining time
                            if last_step != "":
                                passed_steps = float(last_step.split("/")[0])
                                total_steps = float(last_step.split("/")[1])
                                ratio = (total_steps - passed_steps) / passed_steps
                                expected_finish = multiply_time_str(jobs[job_index][3], ratio)
                                # print(expected_finish)
                                job_info += " -> Expected remaining time: %s" % expected_finish

                        # Verify filters
                        if not verify_filters(filters_dict, filled_info_dict):
                            continue

                        _maybe_print("*****************************************")
                        info = colored("(*** Trained ***) ", "magenta") if filled_info_dict[
                                                                               "is_trained"] == "True" else ""
                        _maybe_print(info + full_folder)
                        _maybe_print("*******************************")

                        if "task" in filled_info_dict.keys():
                            filled_info_dict["task"] = decode_task(filled_info_dict["task"])

                        _maybe_print(dict_to_string(filled_info_dict))

                        test_bleu_list = bleu_stats(full_folder + "checkpoints/", ext="tbleu")
                        test_pplx_list = pplx_stats(full_folder + "checkpoints/", ext=".stats")
                        dev_bleu_list = bleu_stats(full_folder + "checkpoints/", ext="dbleu")
                        dev_pplx_list = pplx_stats(full_folder + "checkpoints/", ext=".dstats")

                        _maybe_print(
                            stats_to_string(min_pplx, best_epoch, last_epoch, last_step, job_info, test_bleu_list,
                                            test_pplx_list,
                                            dev_bleu_list, dev_pplx_list))

                        filled_info_dict["folder"] = full_folder
                        filled_info_dict["test_BLEU"] = str(test_bleu_list[0]) if len(test_bleu_list) > 0 else '-'
                        filled_info_dict["test_PPLX"] = str(test_pplx_list[0]) if len(test_pplx_list) > 0 else '-'
                        filled_info_dict["dev_BLEU"] = str(dev_bleu_list[0]) if len(dev_bleu_list) > 0 else '-'
                        filled_info_dict["dev_PPLX"] = str(dev_pplx_list[0]) if len(dev_pplx_list) > 0 else '-'

                        filled_info_dict["min_dev_PPLX_training"] = str(min_pplx)
                        filled_info_dict["epoch(best)"] = str(best_epoch)
                        filled_info_dict["epoch(last)"] = str(last_epoch)
                        filled_info_dict["PPLX history"] = str(pplx_hist)

                        if line == 1:
                            counter = 1
                            for k, v in filled_info_dict.items():
                                ws.cell(row=line, column=counter, value=k)
                                counter += 1
                        counter = 1
                        line += 1
                        for k, v in filled_info_dict.items():
                            ws.cell(row=line, column=counter, value=v)
                            counter += 1

                    except Exception as e:
                        print(full_folder)
                        print("An error occurred while reading the info")
                        print(e)
                        print("------------------------------------")

    wb.save('results.xlsx')
